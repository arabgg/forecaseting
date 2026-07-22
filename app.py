"""
=============================================================================
 DASHBOARD FORECASTING PERSEDIAAN BAHAN BAKU — UMKM F&B (Kendalsari)
 Metode: Moving Average (MA) & Single Exponential Smoothing (SES)
 Stack  : Streamlit · Pandas · Plotly · OpenPyXL
 Periode: November 2025 – Juni 2026 (data nyata dari SO Harian)
=============================================================================
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io, warnings
from datetime import date

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────
#  0. KONFIGURASI HALAMAN
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Forecasting Persediaan | Kendalsari",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────
#  CUSTOM CSS
# ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;600&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif;}
.stApp{background:#0d1117;color:#e6edf3;}
[data-testid="stSidebar"]{background:#161b22!important;border-right:1px solid #21262d;}
[data-testid="stSidebar"] *{color:#c9d1d9!important;}
.kpi{background:linear-gradient(135deg,#161b22,#1c2128);border:1px solid #30363d;
     border-radius:12px;padding:18px 20px;text-align:center;height:100%;}
.kpi-lbl{font-size:.7rem;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
         color:#6e7681;margin-bottom:6px;}
.kpi-val{font-size:2rem;font-weight:700;font-family:'JetBrains Mono',monospace;
         color:#e6edf3;line-height:1;margin-bottom:4px;}
.kpi-sub{font-size:.75rem;color:#484f58;}
.tag-ok{display:inline-block;background:rgba(63,185,80,.15);color:#3fb950;
        border:1px solid rgba(63,185,80,.4);border-radius:20px;padding:3px 14px;
        font-size:.8rem;font-weight:600;}
.tag-err{display:inline-block;background:rgba(248,81,73,.15);color:#f85149;
         border:1px solid rgba(248,81,73,.4);border-radius:20px;padding:3px 14px;
         font-size:.8rem;font-weight:600;animation:blink 1.4s infinite;}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.5}}
.sec{display:flex;align-items:center;gap:8px;margin:26px 0 14px;
     padding-bottom:8px;border-bottom:1px solid #21262d;}
.sec h3{font-size:.8rem;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
        color:#6e7681;margin:0;}
.badge{background:#21262d;color:#6e7681;border-radius:5px;padding:1px 7px;
       font-size:.68rem;font-family:'JetBrains Mono',monospace;}
.mbox{background:#161b22;border:1px solid #30363d;border-radius:10px;
      padding:14px 18px;margin-bottom:8px;}
.mbox-t{font-size:.68rem;font-weight:700;letter-spacing:.11em;text-transform:uppercase;
        color:#484f58;margin-bottom:3px;}
.mbox-v{font-size:1.3rem;font-weight:700;font-family:'JetBrains Mono',monospace;color:#e6edf3;}
.best-box{background:linear-gradient(135deg,rgba(63,185,80,.1),rgba(56,139,253,.07));
          border:1px solid rgba(63,185,80,.3);border-left:4px solid #3fb950;
          border-radius:10px;padding:16px 20px;margin:14px 0;}
.logo-box{background:linear-gradient(135deg,#161b22,#21262d);border-radius:10px;
          padding:14px;text-align:center;margin-bottom:16px;border:1px solid #30363d;}
.logo-box h2{font-size:1rem;font-weight:700;color:#e6edf3;margin:6px 0 2px;}
.logo-box p{font-size:.7rem;color:#484f58;margin:0;}
hr{border-color:#21262d!important;}
label{color:#8b949e!important;font-size:.8rem!important;font-weight:500!important;}
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
#  1. META DATA BAHAN BAKU (dari SO Harian Kendalsari asli)
#     satuan & batas (safety stock) diambil langsung dari Excel
# ═══════════════════════════════════════════════════════════════
ITEM_META = {
    "Angkak":               {"satuan": "gr",     "batas": 1000.0},
    "Ayam Paprika":         {"satuan": "pack",   "batas": 25.0},
    "Bawang Goreng":        {"satuan": "gr",     "batas": 100.0},
    "Bawang Merah":         {"satuan": "gr",     "batas": 350.0},
    "Bawang Putih":         {"satuan": "gr",     "batas": 500.0},
    "Bayam":                {"satuan": "ikat",   "batas": 5.0},
    "Bayam Kremesan":       {"satuan": "Toples", "batas": 10.0},
    "Beef":                 {"satuan": "pack",   "batas": 15.0},
    "Beras":                {"satuan": "kg",     "batas": 3.0},
    "Biang Teh":            {"satuan": "gr",     "batas": 250.0},
    "Boncabe":              {"satuan": "gr",     "batas": 50.0},
    "Bubuk Jelly":          {"satuan": "pack",   "batas": 2.0},
    "Bubuk Pedas":          {"satuan": "Toples", "batas": 2.0},
    "Bumbu Racik Ayam":     {"satuan": "pack",   "batas": 3.0},
    "Cabai Keriting Merah": {"satuan": "gr",     "batas": 250.0},
    "Cabe Ijo Besar":       {"satuan": "gr",     "batas": 250.0},
    "Cabe Merah Besar":     {"satuan": "gr",     "batas": 250.0},
    "Cabe Rawit Ijo":       {"satuan": "gr",     "batas": 150.0},
    "Cabe Rawit Merah":     {"satuan": "gr",     "batas": 150.0},
    "Cup Plastik Takjil":   {"satuan": "pcs",    "batas": 50.0},
    "Cup plastik":          {"satuan": "pcs",    "batas": 50.0},
    "Daun Kemangi":         {"satuan": "ikat",   "batas": 5.0},
    "Es Batu":              {"satuan": "kg",     "batas": 5.0},
    "Galon":                {"satuan": "galon",  "batas": 2.0},
    "Garam":                {"satuan": "gr",     "batas": 100.0},
    "Gas Lpg":              {"satuan": "tabung", "batas": 2.0},
    "Gula Jadi":            {"satuan": "gr",     "batas": 500.0},
    "Gula Pasir":           {"satuan": "gr",     "batas": 1000.0},
    "Handglove":            {"satuan": "pcs",    "batas": 10.0},
    "Jahe Bubuk":           {"satuan": "gr",     "batas": 50.0},
    "Jeruk Limau":          {"satuan": "buah",   "batas": 10.0},
    "Jukut":                {"satuan": "gr",     "batas": 100.0},
    "Kecap Manis":          {"satuan": "ml",     "batas": 100.0},
    "Kertas Plating":       {"satuan": "lembar", "batas": 50.0},
    "Kertas Thermal":       {"satuan": "roll",   "batas": 2.0},
    "Kotak Take away":      {"satuan": "pcs",    "batas": 20.0},
    "Kresek besar":         {"satuan": "pcs",    "batas": 10.0},
    "Kresek kecil":         {"satuan": "pcs",    "batas": 20.0},
    "Kresek tanggung":      {"satuan": "pcs",    "batas": 20.0},
    "Kulit Ayam":           {"satuan": "pack",   "batas": 20.0},
    "Kulit Ayam Paprika":   {"satuan": "pack",   "batas": 10.0},
    "Kunyit Bubuk":         {"satuan": "gr",     "batas": 50.0},
    "Madu":                 {"satuan": "ml",     "batas": 100.0},
    "Merica Bubuk":         {"satuan": "gr",     "batas": 100.0},
    "Micin":                {"satuan": "gr",     "batas": 50.0},
    "Minyak Goreng":        {"satuan": "ml",     "batas": 500.0},
    "Minyak Wijen":         {"satuan": "ml",     "batas": 100.0},
    "Nasi Kemangi":         {"satuan": "pack",   "batas": 10.0},
    "Olesan Sate":          {"satuan": "gr",     "batas": 100.0},
    "Paprika Powder":       {"satuan": "gr",     "batas": 1000.0},
    "Plastik 1 kilo":       {"satuan": "pcs",    "batas": 20.0},
    "Plastik 1/2 kilo":     {"satuan": "pcs",    "batas": 20.0},
    "Plastik 1/4 kilo":     {"satuan": "pcs",    "batas": 20.0},
    "Raja rasa":            {"satuan": "ml",     "batas": 100.0},
    "Roti":                 {"satuan": "buah",   "batas": 5.0},
    "Sabun cuci piring":    {"satuan": "ml",     "batas": 200.0},
    "Sambel Ijo":           {"satuan": "gr",     "batas": 200.0},
    "Sambel Kemangi":       {"satuan": "gr",     "batas": 200.0},
    "Sambel Korek":         {"satuan": "gr",     "batas": 200.0},
    "Sayap Paprika":        {"satuan": "pack",   "batas": 10.0},
    "Sayap Woku":           {"satuan": "pack",   "batas": 10.0},
    "Sedotan":              {"satuan": "pcs",    "batas": 50.0},
    "Serai":                {"satuan": "btg",    "batas": 5.0},
    "Spork (sendok)":       {"satuan": "pcs",    "batas": 30.0},
    "Susu Evaporasi":       {"satuan": "ml",     "batas": 100.0},
    "Susu Kental Manis":    {"satuan": "ml",     "batas": 100.0},
    "Syrup Leci":           {"satuan": "ml",     "batas": 250.0},
    "Syrup Lemon":          {"satuan": "ml",     "batas": 250.0},
    "Syrup Mangga":         {"satuan": "ml",     "batas": 250.0},
    "Syrup Nipis":          {"satuan": "ml",     "batas": 250.0},
    "Syrup Strawberry":     {"satuan": "ml",     "batas": 250.0},
    "Syrup Takjil":         {"satuan": "ml",     "batas": 250.0},
    "T. Maizena":           {"satuan": "gr",     "batas": 100.0},
    "Telur":                {"satuan": "butir",  "batas": 10.0},
    "Terasi":               {"satuan": "gr",     "batas": 50.0},
    "Timun":                {"satuan": "buah",   "batas": 5.0},
    "Tissue":               {"satuan": "pcs",    "batas": 10.0},
    "Tomat Ijo":            {"satuan": "gr",     "batas": 200.0},
    "Tomat Merah":          {"satuan": "gr",     "batas": 1000.0},
    "Tongkol Suwir":        {"satuan": "pack",   "batas": 15.0},
    "Trashbag 45 L":        {"satuan": "pcs",    "batas": 5.0},
    "Trashbag 60 L":        {"satuan": "pcs",    "batas": 5.0},
    "Trashbag 70 L":        {"satuan": "pcs",    "batas": 5.0},
    "UHT":                  {"satuan": "ml",     "batas": 250.0},
    "VIT (air mineral)":    {"satuan": "botol",  "batas": 10.0},
}


# ═══════════════════════════════════════════════════════════════
#  2. MOCK DATA GENERATOR
#     Data harian DENGAN tanggal bolong (sengaja) untuk menguji
#     fungsi reindex + fillna(0). Periode: Nov 2025 – Jun 2026.
# ═══════════════════════════════════════════════════════════════
@st.cache_data
def generate_mock_data() -> pd.DataFrame:
    """
    Simulasi SO Harian 7 bahan baku utama dengan pola realistis.
    Beberapa tanggal sengaja dihapus (warung tutup / tidak ada transaksi)
    agar fungsi pengisian nilai 0 otomatis dapat diuji.
    """
    np.random.seed(2025)
    dates_full = pd.date_range("2025-11-14", "2026-06-28", freq="D")
    n = len(dates_full)

    def weekly_peak(dates_arr):
        return np.array([1.35 if d.weekday() in (4, 5) else 1.0 for d in dates_arr])

    def make_series(base, noise_ratio=0.18, trend=0):
        t  = np.linspace(0, trend * n, n)
        wk = weekly_peak(dates_full)
        ns = np.random.normal(0, base * noise_ratio, n)
        return np.clip(np.round((base + t) * wk + ns, 1), 0, None)

    items_sim = {
        "Ayam Paprika"  : make_series(35,  0.20, 0.01),
        "Tongkol Suwir" : make_series(12,  0.22, 0.005),
        "Kulit Ayam"    : make_series(28,  0.18, 0.008),
        "Beef"          : make_series(8,   0.25, 0.003),
        "Beras"         : make_series(5.5, 0.15, 0.002),
        "Minyak Goreng" : make_series(180, 0.12, 0.05),
        "Gula Pasir"    : make_series(320, 0.10, 0.08),
    }

    rows = []
    for i, d in enumerate(dates_full):
        for item, series in items_sim.items():
            rows.append({"Tanggal": d, "Bahan Baku": item, "Pengeluaran": series[i]})

    df = pd.DataFrame(rows)
    df["Tanggal"] = pd.to_datetime(df["Tanggal"])

    # ── Sengaja hapus beberapa tanggal untuk uji fungsi fillna(0) ──
    # Simulasi: warung tutup di hari tertentu & beberapa hari libur nasional
    drop_dates = pd.to_datetime([
        "2025-12-25", "2025-12-26", "2025-12-31",
        "2026-01-01", "2026-01-27", "2026-01-28",
        "2026-02-01", "2026-02-02", "2026-02-03",
        "2026-03-28", "2026-03-29", "2026-03-30",
        "2026-04-10", "2026-04-11",
        "2026-05-01",
    ])
    df = df[~df["Tanggal"].isin(drop_dates)].reset_index(drop=True)
    print(f"[MockData] {len(df)} records, tanggal bolong: {len(drop_dates)} hari dihapus")
    return df


# ═══════════════════════════════════════════════════════════════
#  3. PARSER FILE EXCEL (SO Harian Kendalsari format asli)
# ═══════════════════════════════════════════════════════════════
SKIP_CATEGORIES = {"LAUK","NASI","MINUMAN","BUMBU","PACKAGING","OPERASIONAL","PIC (PENANGGUNG JAWAB)"}

def _parse_one_sheet(ws) -> list:
    """
    Parse satu sheet SO Harian format pivot (baris=item, kolom=tanggal).
    Mengembalikan list of dict {Tanggal, Bahan Baku, Pengeluaran}.
    """
    rows = list(ws.iter_rows(values_only=True))
    # Deteksi baris header (ada kolom bernama 'ITEM')
    header_idx = next(
        (i for i, row in enumerate(rows[:6])
         if row and any(str(c).strip().upper() == "ITEM" for c in row if c)),
        None
    )
    if header_idx is None:
        return []

    header = rows[header_idx]
    # Kumpulkan kolom tanggal (index >= 5)
    date_cols = {}
    for j, val in enumerate(header):
        if j < 5 or val is None:
            continue
        try:
            ts = pd.Timestamp(val) if hasattr(val, "date") else pd.to_datetime(val)
            date_cols[j] = ts
        except Exception:
            pass

    records = []
    for row in rows[header_idx + 1:]:
        if not row or len(row) < 4:
            continue
        item = row[1]
        if not item or not isinstance(item, str):
            continue
        item = item.strip()
        if not item or item.upper() in SKIP_CATEGORIES:
            continue

        for j, ts in date_cols.items():
            if j >= len(row):
                continue
            raw = row[j]
            if raw is None:
                continue
            # Evaluasi formula Excel sederhana seperti =850/50
            if isinstance(raw, str):
                raw = raw.strip()
                if raw.upper() in ("PACK", "#N/A", "N/A", ""):
                    continue
                if raw.startswith("="):
                    try:
                        raw = eval(raw[1:])
                    except Exception:
                        continue
            try:
                qty = float(raw)
                records.append({
                    "Tanggal"    : ts,
                    "Bahan Baku" : item,
                    "Pengeluaran": max(0.0, qty),
                })
            except Exception:
                continue
    return records


@st.cache_data(show_spinner=False)
def load_excel_file(file_bytes: bytes) -> tuple[pd.DataFrame, str]:
    """
    Baca file Excel SO Harian Kendalsari (semua sheet sekaligus).
    Return: (DataFrame, pesan_status)
    """
    from openpyxl import load_workbook
    try:
        wb    = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        all_r = []
        MONTH_SHEETS = {
            "november 2025","desember 2025","januari 2026",
            "maret 2026","april 2026","mei 2026","juni 2026",
        }
        parsed_sheets = []
        for sh in wb.sheetnames:
            if sh.lower() in MONTH_SHEETS:
                recs = _parse_one_sheet(wb[sh])
                all_r.extend(recs)
                parsed_sheets.append(sh)

        if not all_r:
            return None, "❌ Tidak ada sheet bulan yang dikenali dalam file ini."

        df = pd.DataFrame(all_r)
        df["Tanggal"] = pd.to_datetime(df["Tanggal"])
        df = df[df["Tanggal"].between("2025-01-01", "2026-12-31")]
        df = df[df["Bahan Baku"].str.strip() != ""]

        msg = (f"✅ Berhasil membaca **{len(parsed_sheets)} sheet**: "
               f"{', '.join(parsed_sheets)} — "
               f"**{len(df):,} record**, "
               f"**{df['Bahan Baku'].nunique()} bahan baku**.")
        return df, msg
    except Exception as e:
        return None, f"❌ Gagal membaca file: {e}"


# ═══════════════════════════════════════════════════════════════
#  4. PREPROCESSING — TIME SERIES + IMPUTASI DATA KOSONG
# ═══════════════════════════════════════════════════════════════

IMPUTATION_METHODS = {
    "Forward Fill (LOCF)" : "ffill",
}

IMPUTATION_DESC = {
    "Forward Fill (LOCF)": (
        "Mengisi tanggal kosong dengan <b>nilai terakhir yang tercatat</b> sebelum "
        "tanggal tersebut (<em>Last Observation Carried Forward</em>). Cocok saat data "
        "cenderung stabil tanpa fluktuasi besar harian."
    ),
}


def impute_daily(daily: pd.Series, method: str) -> tuple[pd.Series, int]:
    """
    Terapkan imputasi pada deret harian yang sudah di-reindex.
    NaN = tanggal yang tidak ada record (bukan 0 palsu).

    Parameters
    ----------
    daily  : pd.Series harian dengan NaN di tanggal kosong
    method : "linear" | "dow_mean" | "ffill" | "no_imputation"

    Returns
    -------
    (series_imputed, n_missing) — deret yang sudah terisi & jumlah hari yang diimputasi
    """
    n_missing = int(daily.isna().sum())

    if n_missing == 0:
        return daily.fillna(0), 0

    # Jika imputasi dinonaktifkan → isi semua NaN dengan 0
    if method == "no_imputation":
        return daily.fillna(0), n_missing

    if method == "linear":
        # Interpolasi linear; sisa ujung yang masih NaN → forward-fill lalu backward-fill
        imputed = daily.interpolate(method="linear").ffill().bfill()

    elif method == "dow_mean":
        imputed = daily.copy()
        # Hitung rata-rata per hari-dalam-seminggu dari data yang ada
        dow_means = daily.groupby(daily.index.dayofweek).mean()
        for idx in daily[daily.isna()].index:
            dow = idx.dayofweek
            fill_val = dow_means.get(dow, daily.mean())
            # Jika rata-rata DOW juga NaN (tidak ada data sama sekali di hari itu)
            if pd.isna(fill_val):
                fill_val = daily.mean()
            imputed[idx] = fill_val if not pd.isna(fill_val) else 0.0

    else:  # ffill
        imputed = daily.ffill().bfill().fillna(0)

    # Pastikan tidak ada NaN tersisa & nilai tidak negatif
    imputed = imputed.fillna(0).clip(lower=0)
    return imputed, n_missing


def build_timeseries(
    df: pd.DataFrame,
    bahan_baku: str,
    date_start,
    date_end,
    granularity: str,       # "Harian" | "Mingguan" | "Bulanan"
    imputation: str = "linear",  # "linear" | "dow_mean" | "ffill"
) -> tuple[pd.Series, pd.Series, int]:
    """
    Pipeline preprocessing lengkap:

    1. Filter bahan baku & rentang tanggal
    2. Agregasi harian (sum jika duplikat)
    3. reindex ke rentang penuh → NaN untuk hari tanpa record
    4. Imputasi NaN dengan metode pilihan (SEBELUM resample)
    5. resample ke granularitas (Harian / Mingguan / Bulanan)

    Returns
    -------
    (series_imputed, daily_raw_with_nan, n_missing)
      - series_imputed : deret final siap digunakan forecasting
      - daily_raw      : deret harian sebelum imputasi (untuk visualisasi perbandingan)
      - n_missing      : jumlah hari yang diimputasi
    """
    mask  = (df["Bahan Baku"] == bahan_baku)
    mask &= (df["Tanggal"] >= pd.Timestamp(date_start))
    mask &= (df["Tanggal"] <= pd.Timestamp(date_end))
    sub   = df[mask].copy()

    # Agregasi harian
    daily_agg = (
        sub.groupby("Tanggal")["Pengeluaran"]
        .sum()
        .rename("Pengeluaran")
    )

    # Reindex → NaN untuk hari tanpa record
    full_idx  = pd.date_range(date_start, date_end, freq="D")
    daily_raw = daily_agg.reindex(full_idx)   # NaN di hari kosong

    # Imputasi di level harian
    daily_imputed, n_missing = impute_daily(daily_raw.copy(), imputation)
    daily_imputed = daily_imputed.asfreq("D")

    # Resample sesuai granularitas
    if granularity == "Mingguan":
        series = daily_imputed.resample("W-MON", closed="left", label="left").sum()
    elif granularity == "Bulanan":
        series = daily_imputed.resample("MS").sum()
    else:
        series = daily_imputed

    return series, daily_raw, n_missing


# ═══════════════════════════════════════════════════════════════
#  5. METODE FORECASTING
# ═══════════════════════════════════════════════════════════════
def forecast_ma(series: pd.Series, window: int) -> pd.Series:
    """
    Moving Average — F_t = mean(A_{t-n} ... A_{t-1})
    Digeser 1 periode ke depan (shift=1).
    """
    return series.rolling(window=window, min_periods=window).mean().shift(1)


def forecast_ses(series: pd.Series, alpha: float) -> pd.Series:
    """
    Single Exponential Smoothing
    F_1 = A_0 (inisialisasi),  F_t = α·A_{t-1} + (1−α)·F_{t-1}
    """
    n   = len(series)
    f   = np.full(n, np.nan)
    val = series.values
    if n < 2:
        return pd.Series(f, index=series.index)
    f[1] = val[0]
    for t in range(2, n):
        f[t] = alpha * val[t - 1] + (1 - alpha) * f[t - 1]
    return pd.Series(f, index=series.index)


def next_period_forecast(series: pd.Series, method: str,
                          window: int = 3, alpha: float = 0.3) -> float:
    """Hitung satu nilai forecast untuk t+1."""
    if method == "MA":
        tail = series.dropna().iloc[-window:]
        return float(tail.mean()) if len(tail) >= 1 else 0.0
    # SES
    vals = series.dropna().values
    f    = float(vals[0]) if len(vals) > 0 else 0.0
    for v in vals[1:]:
        f = alpha * v + (1 - alpha) * f
    return f


# ═══════════════════════════════════════════════════════════════
#  6. METRIK EVALUASI AKURASI
# ═══════════════════════════════════════════════════════════════
def _mask(a: np.ndarray, f: np.ndarray):
    return ~np.isnan(f) & ~np.isnan(a) & (a > 0)

def calc_mad(a, f):
    m = _mask(a, f)
    return float(np.mean(np.abs(a[m] - f[m]))) if m.sum() > 0 else np.nan

def calc_mse(a, f):
    m = _mask(a, f)
    return float(np.mean((a[m] - f[m]) ** 2)) if m.sum() > 0 else np.nan

def calc_mape(a, f):
    m = _mask(a, f)
    return float(np.mean(np.abs((a[m] - f[m]) / a[m])) * 100) if m.sum() > 0 else np.nan

def metrics(actual: pd.Series, forecast: pd.Series) -> dict:
    a, f = actual.values.astype(float), forecast.values.astype(float)
    return {"MAD": calc_mad(a, f), "MSE": calc_mse(a, f), "MAPE": calc_mape(a, f)}


# ═══════════════════════════════════════════════════════════════
#  7. VISUALISASI
# ═══════════════════════════════════════════════════════════════
C = {"act": "#58a6ff", "ma": "#f0883e", "ses": "#bc8cff",
     "grid": "#1c2128", "bg": "#0d1117", "font": "#c9d1d9"}

def chart_forecast(series, fc_ma, fc_ses, bahan_baku, window, alpha, granularity):
    lbl_x = {"Harian": "%d %b %Y", "Mingguan": "Minggu %d %b", "Bulanan": "%b %Y"}
    fig   = go.Figure()
    fig.add_trace(go.Scatter(
        x=series.index, y=series.values, name="Aktual",
        mode="lines+markers", line=dict(color=C["act"], width=2),
        marker=dict(size=3),
        hovertemplate="<b>Aktual</b> %{x|" + lbl_x[granularity] + "}: <b>%{y:.2f}</b><extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=fc_ma.index, y=fc_ma.values, name=f"MA (n={window})",
        mode="lines", line=dict(color=C["ma"], width=2, dash="dash"),
        hovertemplate="<b>MA</b> %{x|" + lbl_x[granularity] + "}: <b>%{y:.2f}</b><extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=fc_ses.index, y=fc_ses.values, name=f"SES (α={alpha})",
        mode="lines", line=dict(color=C["ses"], width=2, dash="dot"),
        hovertemplate="<b>SES</b> %{x|" + lbl_x[granularity] + "}: <b>%{y:.2f}</b><extra></extra>",
    ))
    unit   = ITEM_META.get(bahan_baku, {}).get("satuan", "")
    gran_y = {"Harian": f"Pengeluaran/hari ({unit})",
              "Mingguan": f"Pengeluaran/minggu ({unit})",
              "Bulanan": f"Pengeluaran/bulan ({unit})"}
    fig.update_layout(
        title=dict(text=f"<b>{bahan_baku}</b> — Forecasting {granularity}",
                   font=dict(size=14, color=C["font"]), x=0.01),
        xaxis=dict(gridcolor=C["grid"], color=C["font"],
                   rangeslider=dict(visible=True, thickness=0.04)),
        yaxis=dict(title=gran_y[granularity], gridcolor=C["grid"], color=C["font"]),
        legend=dict(orientation="h", y=1.04, x=0,
                    font=dict(color=C["font"], size=11), bgcolor="rgba(0,0,0,0)"),
        plot_bgcolor=C["bg"], paper_bgcolor=C["bg"],
        font=dict(color=C["font"]), hovermode="x unified",
        margin=dict(l=0, r=0, t=55, b=0), height=400,
    )
    return fig


def chart_error_bar(m_ma, m_ses):
    keys  = ["MAD", "MSE", "MAPE"]
    v_ma  = [m_ma.get(k) or 0 for k in keys]
    v_ses = [m_ses.get(k) or 0 for k in keys]
    fig   = go.Figure([
        go.Bar(name="Moving Average", x=keys, y=v_ma,
               marker_color=C["ma"], opacity=.85,
               text=[f"{v:.3f}" for v in v_ma], textposition="outside",
               textfont=dict(color=C["font"])),
        go.Bar(name="SES",            x=keys, y=v_ses,
               marker_color=C["ses"], opacity=.85,
               text=[f"{v:.3f}" for v in v_ses], textposition="outside",
               textfont=dict(color=C["font"])),
    ])
    fig.update_layout(
        barmode="group",
        title=dict(text="<b>Perbandingan Metrik Akurasi</b>",
                   font=dict(size=13, color=C["font"]), x=0.01),
        plot_bgcolor=C["bg"], paper_bgcolor=C["bg"],
        font=dict(color=C["font"]),
        xaxis=dict(color=C["font"], gridcolor=C["grid"]),
        yaxis=dict(color=C["font"], gridcolor=C["grid"]),
        legend=dict(orientation="h", y=1.06, x=0, bgcolor="rgba(0,0,0,0)",
                    font=dict(color=C["font"])),
        margin=dict(l=0, r=0, t=46, b=0), height=260,
    )
    return fig


def chart_imputation(daily_raw: pd.Series, daily_imputed: pd.Series,
                     bahan_baku: str, method_label: str) -> go.Figure:
    """
    Grafik perbandingan data SEBELUM vs SESUDAH imputasi (level harian).
    Titik merah = hari yang diimputasi (semula NaN).
    """
    missing_idx    = daily_raw[daily_raw.isna()].index
    imputed_vals   = daily_imputed.reindex(missing_idx)

    fig = go.Figure()

    # Garis data asli (NaN otomatis terputus → terlihat celah/gap)
    fig.add_trace(go.Scatter(
        x=daily_raw.index, y=daily_raw.values,
        name="Data Asli (ada gap)", mode="lines",
        line=dict(color=C["act"], width=1.5),
        hovertemplate="<b>Asli</b> %{x|%d %b %Y}: %{y:.2f}<extra></extra>",
    ))

    # Garis data setelah imputasi
    fig.add_trace(go.Scatter(
        x=daily_imputed.index, y=daily_imputed.values,
        name=f"Setelah Imputasi ({method_label})", mode="lines",
        line=dict(color="#3fb950", width=1.5, dash="dot"),
        hovertemplate="<b>Imputasi</b> %{x|%d %b %Y}: %{y:.2f}<extra></extra>",
    ))

    # Titik merah = hari yang diimputasi
    if len(missing_idx) > 0:
        fig.add_trace(go.Scatter(
            x=missing_idx, y=imputed_vals.values,
            name="Titik Diimputasi", mode="markers",
            marker=dict(color="#f85149", size=7, symbol="circle",
                        line=dict(color="#ff7b72", width=1)),
            hovertemplate="<b>Diimputasi</b> %{x|%d %b %Y}: %{y:.2f}<extra></extra>",
        ))

    unit = ITEM_META.get(bahan_baku, {}).get("satuan", "")
    fig.update_layout(
        title=dict(
            text=f"<b>Imputasi Data Kosong — {bahan_baku}</b> "
                 f"<span style='font-size:11px;color:#6e7681'>({len(missing_idx)} hari diisi · metode: {method_label})</span>",
            font=dict(size=13, color=C["font"]), x=0.01,
        ),
        xaxis=dict(gridcolor=C["grid"], color=C["font"], tickformat="%d %b %Y"),
        yaxis=dict(title=f"Pengeluaran/hari ({unit})", gridcolor=C["grid"], color=C["font"]),
        legend=dict(orientation="h", y=1.05, x=0, bgcolor="rgba(0,0,0,0)",
                    font=dict(color=C["font"], size=11)),
        plot_bgcolor=C["bg"], paper_bgcolor=C["bg"],
        font=dict(color=C["font"]), hovermode="x unified",
        margin=dict(l=0, r=0, t=55, b=0), height=320,
    )
    return fig


# ═══════════════════════════════════════════════════════════════
#  8. EXPORT EXCEL
# ═══════════════════════════════════════════════════════════════
def build_excel_export(series, fc_ma, fc_ses, m_ma, m_ses,
                       bahan_baku, window, alpha,
                       granularity, best_method,
                       f_ma_t1, f_ses_t1, restock_qty,
                       satuan, imputation_label="Linear Interpolation",
                       n_missing=0) -> bytes:
    """
    Buat file Excel multi-sheet berisi:
      Sheet 1 — Data & Forecast  : tabel lengkap aktual + forecast + error per periode
      Sheet 2 — Metrik Akurasi   : MAD, MSE, MAPE kedua metode + perbandingan
      Sheet 3 — Rekomendasi      : ringkasan parameter, metode terbaik, restock
      Sheet 4 — Info Imputasi    : detail preprocessing data kosong
    """
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference

    buf = io.BytesIO()

    # ── Siapkan data ──
    fmt_map  = {"Harian": "%d %b %Y", "Mingguan": "%d %b %Y", "Bulanan": "%b %Y"}
    fmt      = fmt_map.get(granularity, "%d %b %Y")
    label_t1 = {"Harian": "Besok", "Mingguan": "Minggu Depan", "Bulanan": "Bulan Depan"}
    t1_label = label_t1.get(granularity, "t+1")

    def _v(v):
        return round(float(v), 4) if v is not None and not np.isnan(v) else "N/A"
    def _pct(v):
        return f"{v:.2f}%" if v is not None and not np.isnan(v) else "N/A"

    act  = series.values.astype(float)
    fma  = fc_ma.values.astype(float)
    fses = fc_ses.values.astype(float)

    with pd.ExcelWriter(buf, engine="openpyxl") as xw:

        # ══════════════════════════════════════════
        #  SHEET 1 — Data & Forecast
        # ══════════════════════════════════════════
        df_s1 = pd.DataFrame({
            "Periode"                   : series.index.strftime(fmt),
            "Aktual"                    : np.round(act, 2),
            f"Forecast MA (n={window})" : np.round(fma, 3),
            f"Forecast SES (α={alpha})" : np.round(fses, 3),
            "Error MA (Aktual−MA)"      : np.round(act - fma, 3),
            "Error SES (Aktual−SES)"    : np.round(act - fses, 3),
            "|Error MA|"                : np.round(np.abs(act - fma), 3),
            "|Error SES|"               : np.round(np.abs(act - fses), 3),
        })
        df_s1.to_excel(xw, sheet_name="Data & Forecast", index=False, startrow=2)

        ws1 = xw.sheets["Data & Forecast"]

        # Header judul
        ws1["A1"] = f"Data & Forecast Harian — {bahan_baku} ({granularity})"
        ws1["A1"].font      = Font(bold=True, size=13, color="E6EDF3")
        ws1["A1"].fill      = PatternFill("solid", fgColor="161B22")
        ws1["A2"] = f"Periode: {series.index[0].strftime('%d %b %Y')} — {series.index[-1].strftime('%d %b %Y')}   |   Satuan: {satuan}   |   Imputasi: {imputation_label}"
        ws1["A2"].font      = Font(size=10, color="8B949E", italic=True)

        # Style header kolom (baris 3)
        header_fill  = PatternFill("solid", fgColor="21262D")
        header_font  = Font(bold=True, color="C9D1D9", size=10)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border  = Border(
            bottom=Side(style="thin", color="30363D"),
            right =Side(style="thin", color="30363D"),
        )
        for col_idx, col_name in enumerate(df_s1.columns, start=1):
            cell = ws1.cell(row=3, column=col_idx)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = thin_border

        # Color-code kolom MA (orange) dan SES (purple)
        ma_col_fill  = PatternFill("solid", fgColor="2D2010")
        ses_col_fill = PatternFill("solid", fgColor="1E1428")
        for row in ws1.iter_rows(min_row=4, max_row=ws1.max_row,
                                  min_col=3, max_col=3):
            for cell in row:
                cell.fill = ma_col_fill
        for row in ws1.iter_rows(min_row=4, max_row=ws1.max_row,
                                  min_col=4, max_col=4):
            for cell in row:
                cell.fill = ses_col_fill

        # Auto-width kolom
        for i, col in enumerate(df_s1.columns, start=1):
            max_len = max(len(str(col)), 12)
            ws1.column_dimensions[get_column_letter(i)].width = max_len + 4

        ws1.freeze_panes = "A4"   # freeze judul + header

        # ══════════════════════════════════════════
        #  SHEET 2 — Metrik Akurasi
        # ══════════════════════════════════════════
        rows_met = [
            ["Metrik",  "Rumus",                          "Moving Average", "SES",       "Lebih Baik"],
            ["MAD",     "mean(|Aktual − Forecast|)",      _v(m_ma["MAD"]),  _v(m_ses["MAD"]),  ""],
            ["MSE",     "mean((Aktual − Forecast)²)",     _v(m_ma["MSE"]),  _v(m_ses["MSE"]),  ""],
            ["MAPE (%)", "mean(|Aktual−Forecast|/Aktual)×100", _v(m_ma["MAPE"]), _v(m_ses["MAPE"]), ""],
        ]

        # Tentukan kolom lebih baik (nilai lebih kecil = lebih baik)
        for i, key in enumerate(["MAD", "MSE", "MAPE"], start=1):
            vma  = m_ma.get(key)  or 9999
            vses = m_ses.get(key) or 9999
            rows_met[i][4] = "✅ MA" if vma <= vses else "✅ SES"

        df_s2 = pd.DataFrame(rows_met[1:], columns=rows_met[0])
        df_s2.to_excel(xw, sheet_name="Metrik Akurasi", index=False, startrow=3)

        ws2 = xw.sheets["Metrik Akurasi"]
        ws2["A1"] = f"Evaluasi Akurasi Forecasting — {bahan_baku}"
        ws2["A1"].font = Font(bold=True, size=13, color="E6EDF3")
        ws2["A1"].fill = PatternFill("solid", fgColor="161B22")
        ws2["A2"] = f"Granularitas: {granularity}   |   MA window: n={window}   |   SES α={alpha}   |   Metode Terbaik: {best_method}"
        ws2["A2"].font = Font(size=10, italic=True, color="8B949E")

        # Style header
        for col_idx in range(1, 6):
            cell = ws2.cell(row=4, column=col_idx)
            cell.fill      = PatternFill("solid", fgColor="0D1117")
            cell.font      = Font(bold=True, color="58A6FF", size=11)
            cell.alignment = center_align

        # Color-code baris MA vs SES
        ma_fill2  = PatternFill("solid", fgColor="2D2010")
        ses_fill2 = PatternFill("solid", fgColor="1E1428")
        for row in ws2.iter_rows(min_row=5, max_row=7):
            for cell in row:
                if cell.column == 3:
                    cell.fill = ma_fill2
                elif cell.column == 4:
                    cell.fill = ses_fill2
                elif cell.column == 5:
                    cell.font = Font(bold=True, color="3FB950")

        # Summary box di bawah tabel
        ws2.cell(row=9, column=1).value  = "KESIMPULAN OTOMATIS"
        ws2.cell(row=9, column=1).font   = Font(bold=True, size=11, color="3FB950")
        ws2.cell(row=10, column=1).value = (
            f"Berdasarkan analisis peramalan {granularity.upper()}, metode terbaik untuk "
            f"{bahan_baku} adalah {best_method} dengan MAPE terkecil "
            f"({_pct(min(m_ma.get('MAPE',9999), m_ses.get('MAPE',9999)))})."
        )
        ws2.cell(row=10, column=1).font      = Font(size=10, color="C9D1D9")
        ws2.cell(row=10, column=1).alignment = Alignment(wrap_text=True)
        ws2.merge_cells("A10:E10")
        ws2.row_dimensions[10].height = 40

        for i in range(1, 6):
            ws2.column_dimensions[get_column_letter(i)].width = [16, 36, 18, 18, 14][i-1]

        # ══════════════════════════════════════════
        #  SHEET 3 — Rekomendasi
        # ══════════════════════════════════════════
        param_str = f"(n={window})" if best_method == "Moving Average" else f"(α={alpha})"
        rows_rek = [
            ["PARAMETER",         "NILAI"],
            ["Bahan Baku",        bahan_baku],
            ["Satuan",            satuan],
            ["Granularitas",      granularity],
            ["Periode Analisis",  f"{series.index[0].strftime('%d %b %Y')} – {series.index[-1].strftime('%d %b %Y')}"],
            ["Metode Imputasi",   imputation_label],
            ["Hari Diimputasi",   str(n_missing)],
            ["",                  ""],
            ["HASIL FORECASTING", ""],
            [f"Forecast MA (n={window}) — {t1_label}", f"{f_ma_t1:.2f} {satuan}"],
            [f"Forecast SES (α={alpha}) — {t1_label}", f"{f_ses_t1:.2f} {satuan}"],
            ["",                  ""],
            ["EVALUASI",          ""],
            ["MAPE Moving Average", _pct(m_ma.get("MAPE"))],
            ["MAPE SES",          _pct(m_ses.get("MAPE"))],
            ["Metode Terbaik",    f"{best_method} {param_str}"],
            ["",                  ""],
            ["REKOMENDASI STOK",  ""],
            ["Safety Stock Minimum", f"{m_ma.get('MAD', 0):.2f} {satuan}"],
            ["Forecast Terbaik",  f"{f_ma_t1 if best_method == 'Moving Average' else f_ses_t1:.2f} {satuan}"],
            ["Rekomendasi Order", f"{restock_qty:.2f} {satuan}" if restock_qty > 0 else "Tidak perlu restock"],
        ]

        ws3 = xw.book.create_sheet("Rekomendasi")
        ws3["A1"] = f"Laporan Rekomendasi Forecasting — {bahan_baku}"
        ws3["A1"].font = Font(bold=True, size=14, color="E6EDF3")
        ws3["A1"].fill = PatternFill("solid", fgColor="0D1117")
        ws3.merge_cells("A1:B1")

        section_keys = {"PARAMETER", "HASIL FORECASTING", "EVALUASI", "REKOMENDASI STOK"}
        for r_idx, (k, v) in enumerate(rows_rek, start=3):
            c_k = ws3.cell(row=r_idx, column=1, value=k)
            c_v = ws3.cell(row=r_idx, column=2, value=v)
            if k in section_keys:
                c_k.font = Font(bold=True, size=10, color="58A6FF")
                c_k.fill = PatternFill("solid", fgColor="161B22")
                c_v.fill = PatternFill("solid", fgColor="161B22")
            elif k == "Metode Terbaik":
                c_k.font = Font(bold=True, color="3FB950")
                c_v.font = Font(bold=True, color="3FB950")
            elif k == "Rekomendasi Order":
                fill_c = "3D0A0A" if restock_qty > 0 else "0A2D0A"
                font_c = "F85149" if restock_qty > 0 else "3FB950"
                c_k.fill = PatternFill("solid", fgColor=fill_c)
                c_v.fill = PatternFill("solid", fgColor=fill_c)
                c_k.font = Font(bold=True, color=font_c, size=11)
                c_v.font = Font(bold=True, color=font_c, size=11)
            else:
                c_k.font = Font(color="8B949E", size=10)
                c_v.font = Font(color="C9D1D9", size=10)

        ws3.column_dimensions["A"].width = 36
        ws3.column_dimensions["B"].width = 30

        # ══════════════════════════════════════════
        #  SHEET 4 — Info Imputasi
        # ══════════════════════════════════════════
        imputation_desc = {
            "Linear Interpolation": (
                "Mengisi tanggal kosong berdasarkan garis tren linear antara "
                "nilai hari sebelum dan sesudahnya (.interpolate(method='linear')). "
                "Cocok saat data memiliki tren naik/turun yang konsisten."
            ),
            "Day-of-Week Mean": (
                "Mengisi tanggal kosong menggunakan rata-rata hari yang sama dalam seminggu. "
                "Misal: Sabtu kosong → diisi rata-rata semua Sabtu yang tersedia. "
                "Sangat cocok untuk bisnis F&B dengan pola weekend peak."
            ),
            "Forward Fill (LOCF)": (
                "Last Observation Carried Forward: mengisi tanggal kosong dengan "
                "nilai terakhir yang tercatat sebelum tanggal tersebut (.ffill()). "
                "Cocok saat data cenderung stabil tanpa fluktuasi besar harian."
            ),
        }

        ws4 = xw.book.create_sheet("Info Imputasi")
        ws4["A1"] = "Laporan Preprocessing & Imputasi Data Kosong"
        ws4["A1"].font = Font(bold=True, size=13, color="E6EDF3")
        ws4["A1"].fill = PatternFill("solid", fgColor="0D1117")
        ws4.merge_cells("A1:C1")

        info_rows = [
            ("Bahan Baku Dianalisis",     bahan_baku),
            ("Total Periode Harian",      str((series.index[-1] - series.index[0]).days + 1)),
            ("Hari Ada Record",           str((series.index[-1] - series.index[0]).days + 1 - n_missing)),
            ("Hari Tanpa Record (NaN)",   str(n_missing)),
            ("Persentase Data Kosong",    f"{n_missing / max((series.index[-1]-series.index[0]).days+1,1)*100:.1f}%"),
            ("Metode Imputasi Dipilih",   imputation_label),
            ("Penjelasan Metode",         imputation_desc.get(imputation_label, "")),
            ("False Zero Bias",           "DIHINDARI — data kosong TIDAK diisi 0 mentah-mentah"),
            ("Catatan Akademis",
             "Pengisian nilai 0 langsung pada hari tanpa transaksi (false zero) "
             "akan menyebabkan bias pada MAPE/MSE dan menghasilkan rekomendasi "
             "restock yang terlalu rendah (risiko stockout). "
             "Imputasi yang tepat menghasilkan deret waktu yang lebih representatif."),
        ]

        for r_idx, (k, v) in enumerate(info_rows, start=3):
            c_k = ws4.cell(row=r_idx, column=1, value=k)
            c_v = ws4.cell(row=r_idx, column=2, value=v)
            c_k.font      = Font(color="8B949E", size=10, bold=True)
            c_v.font      = Font(color="C9D1D9", size=10)
            c_v.alignment = Alignment(wrap_text=True)
            if k in ("Hari Tanpa Record (NaN)", "Metode Imputasi Dipilih", "False Zero Bias"):
                c_k.font = Font(color="58A6FF", bold=True, size=10)
                c_v.font = Font(color="F0883E", bold=True, size=10)
            if k == "Catatan Akademis":
                c_k.font = Font(color="3FB950", bold=True, size=10)
                c_v.font = Font(color="C9D1D9", size=10, italic=True)

        ws4.column_dimensions["A"].width = 32
        ws4.column_dimensions["B"].width = 80
        ws4.row_dimensions[11].height = 60  # baris catatan akademis lebih tinggi

    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
#  9. SIDEBAR
# ═══════════════════════════════════════════════════════════════
def render_sidebar(df: pd.DataFrame):
    with st.sidebar:
        # Logo
        st.markdown("""
        <div class="logo-box">
            <div style="font-size:2rem">📦</div>
            <h2>ForecastIQ</h2>
            <p>Persediaan Bahan Baku<br>UMKM F&amp;B · Kendalsari</p>
        </div>""", unsafe_allow_html=True)

        # ── File Uploader ──
        st.markdown("#### 📂 Upload Data SO Harian")
        uploaded = st.file_uploader(
            "File Excel (.xlsx)",
            type=None,                  # Biarkan None agar semua tipe bisa masuk → kita validasi manual
            label_visibility="collapsed",
            help="Upload file SO Harian Kendalsari format Excel (.xlsx)",
        )

        # ── Peringatan format & link template ──
        st.warning(
            "⚠️ **Pastikan format file sesuai template!**\n\n"
            "File harus berbentuk Excel (.xlsx) dengan sheet bernama bulan "
            "(contoh: *Maret 2026*), kolom NO · ITEM · Satuan · Batas, "
            "dan kolom tanggal mulai kolom ke-6.\n\n"
            "📥 [**Download Template Dokumen**](https://docs.google.com/spreadsheets/d/"
            "1Gw-txJTPWFGI8fXLDGU2AidZtV0vM6j4v4Ef5YrhiLc/edit?usp=sharing)"
        )

        # ── Validasi tipe file & pilih sumber data ──
        use_real = False
        df_real  = None
        if uploaded is not None:
            # Validasi ekstensi file secara manual
            file_name = uploaded.name.lower()
            if not file_name.endswith(".xlsx"):
                st.error(
                    "❌ **Format file tidak valid!**\n\n"
                    f"File **'{uploaded.name}'** bukan file Excel. "
                    "Hanya file berformat **`.xlsx`** yang diterima. "
                    "Silakan upload ulang dengan file yang benar."
                )
                st.info("🎲 Menggunakan **data demo** sebagai gantinya.")
            else:
                with st.spinner("Membaca file..."):
                    file_bytes = uploaded.read()
                    df_real, msg = load_excel_file(file_bytes)
                if df_real is not None:
                    st.success(msg)
                    use_real = True
                    df = df_real
                else:
                    st.error(msg)
                    st.info("Menggunakan **data demo** sebagai gantinya.")

        if not use_real and uploaded is None:
            st.info("🎲 Menggunakan **data demo** (mock data dengan tanggal bolong).")

        st.markdown("---")

        # ── Pilih Bahan Baku ──
        st.markdown("#### 🥩 Bahan Baku")
        items = sorted(df["Bahan Baku"].unique().tolist())
        default_item = "Ayam Paprika" if "Ayam Paprika" in items else items[0]
        bahan_baku = st.selectbox("Pilih bahan baku", items,
                                   index=items.index(default_item))

        st.markdown("---")

        # ── Filter Tanggal ──
        st.markdown("#### 📅 Rentang Tanggal")
        mn = df["Tanggal"].min().date()
        mx = df["Tanggal"].max().date()
        d1 = st.date_input("Dari", value=mn, min_value=mn, max_value=mx)
        d2 = st.date_input("Sampai", value=mx, min_value=mn, max_value=mx)
        if d1 > d2:
            st.error("Tanggal awal > akhir!")
            d1, d2 = mn, mx

        st.markdown("---")

        # ── Granularitas Waktu ──
        st.markdown("#### ⏱ Granularitas Waktu")
        granularity = st.radio(
            "Tampilkan data per:", ["Harian", "Mingguan", "Bulanan"],
            help="Data harian akan di-resample otomatis sesuai pilihan ini",
        )

        st.markdown("---")

        # ── Parameter MA ──
        st.markdown("#### 📊 Moving Average")
        if granularity == "Harian":
            win_label = "Window (n hari)"
        elif granularity == "Mingguan":
            win_label = "Window (n minggu)"
        else:
            win_label = "Window (n bulan)"
        ma_window = st.radio(win_label, [3, 5, 7], index=0)

        # ── Parameter SES ──
        st.markdown("#### 📈 Single Exponential Smoothing")
        alpha = st.slider("Alpha (α)", 0.1, 0.9, 0.3, 0.1)
        st.caption(
            f"α = **{alpha:.1f}** — "
            f"{'Sangat responsif' if alpha >= 0.7 else 'Responsif' if alpha >= 0.5 else 'Stabil'}"
        )

        st.markdown("---")

        # ── Metode Imputasi Data Kosong ──
        st.markdown("#### 🩹 Imputasi Data Kosong")

        # Toggle: aktifkan atau nonaktifkan imputasi (Forward Fill)
        use_imputation = st.toggle(
            "Gunakan Imputasi",
            value=True,
            help=(
                "Aktifkan untuk mengisi tanggal kosong menggunakan Forward Fill (LOCF). "
                "Jika dinonaktifkan, hari tanpa record akan diisi dengan nilai 0."
            ),
        )

        if use_imputation:
            imputation_label = "Forward Fill (LOCF)"
            imputation_key   = "ffill"
            st.markdown(
                '<div style="background:#1c2128;border:1px solid #30363d;border-left:3px solid #58a6ff;'
                'border-radius:8px;padding:10px 12px;font-size:.75rem;color:#8b949e;line-height:1.6">'
                '✅ Imputasi <b style="color:#58a6ff">aktif</b> — metode <b>Forward Fill (LOCF)</b>.<br>'
                'Mengisi tanggal kosong dengan <b>nilai terakhir yang tercatat</b> sebelum '
                'tanggal tersebut (<em>Last Observation Carried Forward</em>).'
                '</div>',
                unsafe_allow_html=True,
            )
        else:
            imputation_label = "Tanpa Imputasi (isi 0)"
            imputation_key   = "no_imputation"
            st.markdown(
                '<div style="background:#1c2128;border:1px solid #30363d;border-left:3px solid #f85149;'
                'border-radius:8px;padding:10px 12px;font-size:.75rem;color:#8b949e;line-height:1.6">'
                '⚠️ Imputasi <b style="color:#f85149">dinonaktifkan</b>. '
                'Semua tanggal tanpa record akan diisi <b>0</b>. '
                'Perhatikan potensi <em>false zero bias</em> pada hasil forecasting.'
                '</div>',
                unsafe_allow_html=True,
            )

        st.markdown("---")
        st.markdown(
            '<div style="text-align:center;color:#21262d;font-size:.68rem">'
            'Skripsi S1 · Sistem Informasi · 2026</div>',
            unsafe_allow_html=True,
        )

    return df, bahan_baku, d1, d2, granularity, ma_window, alpha, imputation_label, imputation_key


# ═══════════════════════════════════════════════════════════════
#  10. MAIN DASHBOARD
# ═══════════════════════════════════════════════════════════════
def render_dashboard(df, bahan_baku, d1, d2, granularity, ma_window, alpha,
                     imputation_label="Linear Interpolation", imputation_key="linear"):

    # ── Header ──
    gran_map = {"Harian": "HARIAN", "Mingguan": "MINGGUAN", "Bulanan": "BULANAN"}
    st.markdown(f"""
    <span style="font-size:.68rem;letter-spacing:.12em;text-transform:uppercase;
                 color:#484f58;font-weight:600">
        DASHBOARD FORECASTING {gran_map[granularity]} — UMKM F&amp;B KENDALSARI
    </span>
    <h1 style="font-size:1.7rem;font-weight:700;color:#e6edf3;margin:2px 0 4px">
        {bahan_baku}
    </h1>
    <p style="color:#484f58;font-size:.8rem;margin:0">
        Periode: <b style="color:#6e7681">{d1.strftime('%d %b %Y')}</b>
        — <b style="color:#6e7681">{d2.strftime('%d %b %Y')}</b>
        &nbsp;·&nbsp; Granularitas: <b style="color:#6e7681">{granularity}</b>
        &nbsp;·&nbsp; Imputasi: <b style="color:#58a6ff">{imputation_label}</b>
    </p>""", unsafe_allow_html=True)

    # ── Build time series dengan imputasi ──
    series, daily_raw, n_missing = build_timeseries(
        df, bahan_baku, d1, d2, granularity, imputation_key
    )
    # Rekonstruksi daily_imputed untuk grafik imputasi (level harian, sebelum resample)
    full_idx      = pd.date_range(d1, d2, freq="D")
    daily_raw_f   = daily_raw.reindex(full_idx) if not daily_raw.index.equals(full_idx) else daily_raw
    daily_imputed_vis, _ = impute_daily(daily_raw_f.copy(), imputation_key)

    if len(series) < 5:
        st.warning("⚠️ Data terlalu sedikit (< 5 periode). Perluas rentang tanggal.")
        return

    n_zero_filled = n_missing   # renamed for clarity below

    # ── Forecasting ──
    fc_ma  = forecast_ma(series, ma_window)
    fc_ses = forecast_ses(series, alpha)

    # ── Metrik ──
    m_ma  = metrics(series, fc_ma)
    m_ses = metrics(series, fc_ses)

    mape_ma  = m_ma["MAPE"]  if m_ma["MAPE"]  and not np.isnan(m_ma["MAPE"])  else 9999
    mape_ses = m_ses["MAPE"] if m_ses["MAPE"] and not np.isnan(m_ses["MAPE"]) else 9999
    best      = "Moving Average" if mape_ma <= mape_ses else "SES"
    best_mape = min(mape_ma, mape_ses)
    best_col  = C["ma"] if best == "Moving Average" else C["ses"]

    # ── Info stok ──
    meta       = ITEM_META.get(bahan_baku, {"satuan": "unit", "batas": 10.0})
    satuan     = meta["satuan"]
    safety     = meta["batas"]
    stok_kini  = float(series.iloc[-1]) if len(series) > 0 else 0.0
    stok_aman  = stok_kini >= safety

    # ── Forecast t+1 ──
    f_ma_t1  = next_period_forecast(series, "MA",  ma_window, alpha)
    f_ses_t1 = next_period_forecast(series, "SES", ma_window, alpha)
    f_best   = f_ma_t1 if best == "Moving Average" else f_ses_t1
    restock  = max(0.0, round(f_best - stok_kini, 2))

    label_t1 = {"Harian": "Besok", "Mingguan": "Minggu Depan", "Bulanan": "Bulan Depan"}

    # ─────────────────────────────────────────────
    #  BAGIAN 1 — KPI CARDS
    # ─────────────────────────────────────────────
    st.markdown("""<div class="sec"><h3>Ringkasan Stok</h3>
        <span class="badge">01</span></div>""", unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f"""<div class="kpi">
            <div class="kpi-lbl">Pengeluaran Terakhir</div>
            <div class="kpi-val">{stok_kini:.1f}</div>
            <div class="kpi-sub">{satuan} / periode terakhir</div>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class="kpi">
            <div class="kpi-lbl">Safety Stock</div>
            <div class="kpi-val">{safety:.1f}</div>
            <div class="kpi-sub">{satuan} (batas minimum)</div>
        </div>""", unsafe_allow_html=True)
    with c3:
        tag = '<span class="tag-ok">✅ AMAN</span>' if stok_aman else '<span class="tag-err">🚨 HARUS ORDER!</span>'
        note = "Stok di atas minimum" if stok_aman else "Stok di bawah safety stock!"
        st.markdown(f"""<div class="kpi">
            <div class="kpi-lbl">Status</div>
            <div style="margin-top:14px">{tag}</div>
            <div class="kpi-sub" style="margin-top:8px">{note}</div>
        </div>""", unsafe_allow_html=True)
    with c4:
        st.markdown(f"""<div class="kpi">
            <div class="kpi-lbl">Periode Data</div>
            <div class="kpi-val" style="font-size:1.3rem">{len(series)}</div>
            <div class="kpi-sub">{granularity.lower()} · {n_zero_filled} hari bolong → 0</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ─────────────────────────────────────────────
    #  PANEL IMPUTASI DATA KOSONG
    # ─────────────────────────────────────────────
    if n_zero_filled > 0:
        imp_color = {"Linear Interpolation": "#58a6ff",
                     "Day-of-Week Mean"    : "#f0883e",
                     "Forward Fill (LOCF)" : "#bc8cff"}.get(imputation_label, "#58a6ff")

        # Badge metode imputasi
        st.markdown(f"""
        <div style="background:#161b22;border:1px solid #30363d;border-left:4px solid {imp_color};
                    border-radius:10px;padding:14px 18px;margin-bottom:14px">
            <div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap">
                <div>
                    <div style="font-size:.68rem;font-weight:700;letter-spacing:.1em;
                                text-transform:uppercase;color:#484f58;margin-bottom:4px">
                        🩹 PREPROCESSING — IMPUTASI DATA KOSONG
                    </div>
                    <div style="font-size:.9rem;color:#e6edf3">
                        Ditemukan <strong style="color:{imp_color}">{n_zero_filled} hari</strong>
                        tanpa record dalam data <strong>{bahan_baku}</strong>.
                        Diisi menggunakan metode
                        <strong style="color:{imp_color}">{imputation_label}</strong>
                        sebelum data di-resample ke level <em>{granularity.lower()}</em>.
                    </div>
                    <div style="font-size:.78rem;color:#6e7681;margin-top:6px">
                        ⚠️ <em>False Zero dihindari</em> — hari tanpa transaksi <strong>tidak</strong>
                        langsung diisi 0 untuk mencegah bias error MAPE yang membengkak.
                    </div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Grafik imputasi (ekspandable)
        with st.expander(
            f"📊 Lihat Grafik Imputasi — {n_zero_filled} hari diisi dengan {imputation_label}",
            expanded=False,
        ):
            st.plotly_chart(
                chart_imputation(daily_raw_f, daily_imputed_vis, bahan_baku, imputation_label),
                use_container_width=True,
            )
            st.caption(
                "🔴 Titik merah = hari yang semula kosong (NaN) dan sudah diimputasi. "
                "Garis biru putus = data asli (celah = hari tidak ada record). "
                "Garis hijau = setelah imputasi."
            )

    else:
        st.success(
            f"✅ **Data lengkap** — tidak ada tanggal kosong dalam rentang "
            f"**{d1.strftime('%d %b %Y')} – {d2.strftime('%d %b %Y')}** untuk **{bahan_baku}**."
        )

    # ─────────────────────────────────────────────
    #  BAGIAN 2 — GRAFIK UTAMA
    # ─────────────────────────────────────────────
    st.markdown("""<div class="sec"><h3>Visualisasi Forecasting</h3>
        <span class="badge">02</span></div>""", unsafe_allow_html=True)
    st.plotly_chart(
        chart_forecast(series, fc_ma, fc_ses, bahan_baku, ma_window, alpha, granularity),
        use_container_width=True,
    )

    # ─────────────────────────────────────────────
    #  BAGIAN 3 — EVALUASI AKURASI
    # ─────────────────────────────────────────────
    st.markdown("""<div class="sec"><h3>Evaluasi Akurasi</h3>
        <span class="badge">03</span></div>""", unsafe_allow_html=True)

    def _mbox(label, value, fmt=".4f"):
        v = f"{value:{fmt}}" if (value is not None and not np.isnan(value)) else "N/A"
        st.markdown(f"""<div class="mbox">
            <div class="mbox-t">{label}</div>
            <div class="mbox-v">{v}</div>
        </div>""", unsafe_allow_html=True)

    col_ma, _, col_ses = st.columns([1, 0.04, 1])
    is_best_ma  = best == "Moving Average"
    is_best_ses = best == "SES"

    with col_ma:
        bc = C["ma"] if is_best_ma else "#30363d"
        bk = f'<span style="font-size:.68rem;background:{C["ma"]}20;color:{C["ma"]};padding:2px 8px;border-radius:12px;border:1px solid {C["ma"]}55">⭐ TERBAIK</span>' if is_best_ma else ""
        st.markdown(f"""<div style="background:#161b22;border:2px solid {bc};border-radius:12px;
            padding:16px 18px;margin-bottom:10px">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
                <span style="font-size:.9rem;font-weight:700;color:{C['ma']}">📊 Moving Average (n={ma_window})</span>
                {bk}
            </div></div>""", unsafe_allow_html=True)
        _mbox("MAD — Mean Absolute Deviation",              m_ma["MAD"])
        _mbox("MSE — Mean Squared Error",                   m_ma["MSE"])
        _mbox("MAPE — Mean Absolute Percentage Error (%)",  m_ma["MAPE"], ".2f")

    with _:
        st.markdown('<div style="border-left:1px solid #21262d;height:100%;margin:0 auto;width:1px"></div>',
                    unsafe_allow_html=True)

    with col_ses:
        bc = C["ses"] if is_best_ses else "#30363d"
        bk = f'<span style="font-size:.68rem;background:{C["ses"]}20;color:{C["ses"]};padding:2px 8px;border-radius:12px;border:1px solid {C["ses"]}55">⭐ TERBAIK</span>' if is_best_ses else ""
        st.markdown(f"""<div style="background:#161b22;border:2px solid {bc};border-radius:12px;
            padding:16px 18px;margin-bottom:10px">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
                <span style="font-size:.9rem;font-weight:700;color:{C['ses']}">📈 Single Exponential Smoothing (α={alpha})</span>
                {bk}
            </div></div>""", unsafe_allow_html=True)
        _mbox("MAD — Mean Absolute Deviation",              m_ses["MAD"])
        _mbox("MSE — Mean Squared Error",                   m_ses["MSE"])
        _mbox("MAPE — Mean Absolute Percentage Error (%)",  m_ses["MAPE"], ".2f")

    st.plotly_chart(chart_error_bar(m_ma, m_ses), use_container_width=True)

    # ── Kesimpulan otomatis (menyesuaikan granularitas) ──
    gran_kalimat = {"Harian": "HARIAN", "Mingguan": "MINGGUAN", "Bulanan": "BULANAN"}
    param_str = f"(n={ma_window})" if best == "Moving Average" else f"(α={alpha})"
    st.markdown(f"""
    <div class="best-box">
        <div style="font-size:.68rem;font-weight:700;letter-spacing:.1em;
                    text-transform:uppercase;color:#484f58;margin-bottom:6px">
            🤖 KESIMPULAN OTOMATIS
        </div>
        <div style="font-size:.95rem;color:#e6edf3;line-height:1.7">
            Berdasarkan analisis peramalan <strong style="color:{best_col}">
            {gran_kalimat[granularity]}</strong>, metode terbaik untuk
            <strong style="color:{best_col}">{bahan_baku}</strong> adalah
            <strong style="color:{best_col}">{best} {param_str}</strong>
            dengan tingkat error MAPE sebesar
            <strong style="color:{best_col}">{best_mape:.2f}%</strong>.
            Metode ini menghasilkan prediksi yang lebih akurat dibandingkan metode lainnya
            berdasarkan data {granularity.lower()} periode
            {d1.strftime('%d %b %Y')} – {d2.strftime('%d %b %Y')}.
        </div>
    </div>""", unsafe_allow_html=True)

    # ─────────────────────────────────────────────
    #  BAGIAN 4 — FORECAST T+1 & RESTOCK
    # ─────────────────────────────────────────────
    st.markdown(f"""<div class="sec">
        <h3>Forecast {label_t1[granularity]} &amp; Rekomendasi Restock</h3>
        <span class="badge">04</span></div>""", unsafe_allow_html=True)

    col_tbl, col_rek = st.columns(2)

    with col_tbl:
        def _mape_str(v):
            return f"{v:.2f}%" if v and not np.isnan(v) else "N/A"
        df_t1 = pd.DataFrame({
            "Metode"                  : ["Moving Average (MA)", f"SES (α={alpha})"],
            f"Forecast {label_t1[granularity]}" : [f"{f_ma_t1:.2f} {satuan}", f"{f_ses_t1:.2f} {satuan}"],
            "MAPE (%)"                : [_mape_str(m_ma["MAPE"]), _mape_str(m_ses["MAPE"])],
            "Status"                  : ["⭐ Terbaik" if best=="Moving Average" else "—",
                                         "⭐ Terbaik" if best=="SES" else "—"],
        })
        st.markdown(f"**📅 Prediksi Kebutuhan — {label_t1[granularity]}**")
        st.dataframe(df_t1, use_container_width=True, hide_index=True)

    with col_rek:
        rc = "#f85149" if restock > 0 else "#3fb950"
        rl = f"{restock:.2f} {satuan}" if restock > 0 else "Tidak perlu restock"
        st.markdown("**📋 Rekomendasi Pembelian**")
        st.markdown(f"""
        <div style="background:#161b22;border:1px solid #30363d;border-radius:10px;padding:18px">
            <table style="width:100%;border-collapse:collapse;font-size:.84rem;color:#c9d1d9">
                <tr style="border-bottom:1px solid #21262d">
                    <td style="padding:7px 4px;color:#484f58">Metode Terbaik</td>
                    <td style="padding:7px 4px;text-align:right;font-weight:600;color:{best_col}">{best} {param_str}</td>
                </tr>
                <tr style="border-bottom:1px solid #21262d">
                    <td style="padding:7px 4px;color:#484f58">Forecast {label_t1[granularity]}</td>
                    <td style="padding:7px 4px;text-align:right;font-weight:600">{f_best:.2f} {satuan}</td>
                </tr>
                <tr style="border-bottom:1px solid #21262d">
                    <td style="padding:7px 4px;color:#484f58">Pengeluaran Periode Terakhir</td>
                    <td style="padding:7px 4px;text-align:right;font-weight:600">{stok_kini:.2f} {satuan}</td>
                </tr>
                <tr style="border-bottom:1px solid #21262d">
                    <td style="padding:7px 4px;color:#484f58">Safety Stock Minimum</td>
                    <td style="padding:7px 4px;text-align:right;font-weight:600">{safety:.2f} {satuan}</td>
                </tr>
                <tr>
                    <td style="padding:9px 4px;font-weight:700;color:#e6edf3">🛒 Rekomendasi Order</td>
                    <td style="padding:9px 4px;text-align:right;font-weight:700;font-size:1.05rem;color:{rc}">{rl}</td>
                </tr>
            </table>
        </div>""", unsafe_allow_html=True)

    # ─────────────────────────────────────────────
    #  BAGIAN 5 — DOWNLOAD EXCEL (menonjol)
    # ─────────────────────────────────────────────
    st.markdown("""<div class="sec"><h3>Export Hasil Forecasting</h3>
        <span class="badge">05</span></div>""", unsafe_allow_html=True)

    xlsx_bytes = build_excel_export(
        series, fc_ma, fc_ses, m_ma, m_ses,
        bahan_baku, ma_window, alpha, granularity, best,
        f_ma_t1, f_ses_t1, restock, satuan,
        imputation_label=imputation_label,
        n_missing=n_zero_filled,
    )
    fname = f"forecast_{bahan_baku.replace(' ','_')}_{granularity}_{d1}_{d2}.xlsx"

    col_dl, col_info = st.columns([1, 2])

    with col_dl:
        st.download_button(
            label="⬇️  Download Hasil Forecasting (.xlsx)",
            data=xlsx_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
        st.caption(f"📄 `{fname}`")

    with col_info:
        st.markdown(f"""
        <div style="background:#161b22;border:1px solid #30363d;border-radius:10px;
                    padding:14px 18px;font-size:.82rem;color:#8b949e;line-height:1.8">
            File Excel berisi <strong style="color:#c9d1d9">4 sheet</strong>:<br>
            &nbsp; 📊 <strong style="color:#f0883e">Data &amp; Forecast</strong>
                — tabel aktual + MA + SES + error per periode<br>
            &nbsp; 📈 <strong style="color:#bc8cff">Metrik Akurasi</strong>
                — MAD, MSE, MAPE kedua metode + kesimpulan<br>
            &nbsp; 📋 <strong style="color:#3fb950">Rekomendasi</strong>
                — parameter, metode terbaik, jumlah order<br>
            &nbsp; 🩹 <strong style="color:#58a6ff">Info Imputasi</strong>
                — detail preprocessing &amp; catatan akademis
        </div>""", unsafe_allow_html=True)

    # ─────────────────────────────────────────────
    #  BAGIAN 6 — TABEL LENGKAP (ekspandable)
    # ─────────────────────────────────────────────
    with st.expander("📂 Lihat Tabel Data Lengkap (Aktual vs Forecast)", expanded=False):
        fmt = {"Harian": "%d %b %Y", "Mingguan": "%d %b %Y", "Bulanan": "%b %Y"}
        df_full = pd.DataFrame({
            "Periode"                      : series.index.strftime(fmt[granularity]),
            "Aktual"                       : series.values.round(2),
            f"Forecast MA (n={ma_window})" : fc_ma.values.round(3),
            f"Forecast SES (α={alpha})"    : fc_ses.values.round(3),
            "Error MA"                     : (series.values - fc_ma.values).round(3),
            "Error SES"                    : (series.values - fc_ses.values).round(3),
        })
        st.dataframe(df_full, use_container_width=True, hide_index=True, height=300)

    # Footer
    st.markdown("---")
    st.markdown("""
    <div style="text-align:center;color:#21262d;font-size:.7rem;padding:6px 0">
        ForecastIQ · UMKM F&amp;B Kendalsari · Moving Average &amp; SES ·
        Built with Streamlit + Plotly + Pandas
    </div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
#  11. ENTRY POINT
# ═══════════════════════════════════════════════════════════════
def main():
    df_mock = generate_mock_data()
    df, bahan_baku, d1, d2, granularity, ma_window, alpha, imputation_label, imputation_key = render_sidebar(df_mock)
    render_dashboard(df, bahan_baku, d1, d2, granularity, ma_window, alpha,
                     imputation_label, imputation_key)

if __name__ == "__main__":
    main()
